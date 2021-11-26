import os
import openpyxl
from Web_Key_demo1 import WebKeyDemo1
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment

pwd_dir = os.path.dirname(os.path.abspath(__file__))
pwd_excel = os.path.join(pwd_dir, 'at_cases.xlsx')
# 读取excel
excel = openpyxl.load_workbook(pwd_excel)
# 读取sheet表
# sh = excel['Sheet1']

# 读取所有sheet
sheets = excel.sheetnames  # 获取所有的sheet表名
for sheet in sheets:  # 遍历表名，将遍历的sheet传给excel[sheet]
    sheet_temp = excel[sheet]
    # 解决第三方库没有代码提示
    assert isinstance(sheet_temp, openpyxl.workbook.workbook.Worksheet)
    for values in sheet_temp.values:  # 将得到的所有sheet加载
        # 只需要获取用例，不需要标题，通过if判断
        if type(values[0]) is int:  # 从角标为0开始判断是否为数字
            '''
            用例结构
                1.编号：不用管
                2.调用的关键字：结合关键字驱动类应用
                3.关键字对应的参数：测试数据，结合关键字驱动类应该
                4.本条用例的行为描述：行为记录，可以添加到日志中
            '''
            # 提取数据，name\value\txt是对应的关键字类的输入
            data = {}
            data['name'] = values[2]
            data['value'] = values[3]
            data['txt'] = values[4]
            # 优化测试数据内容，将所有未None的数据全部从data中清除
            for key in list(data.keys()):
                if data[key] is None:
                    del data[key]
            # print(data)
            # 基于操作行为和对应参数来执行自动化操作
            '''
                用例的操作行为主要分为：
                    1. 实例化，通过一个操作行为实例化关键字驱动类对象
                    2. 常规操作，通过调用已实例化的对象，执行对应的函数。
                    3. 断言操作，判断预期与实际是否符合，将结果填入测试用例中。
            '''
            # 实例化关键字驱动
            if values[1] == 'open_browser':
                keys = WebKeyDemo1(**data)
                # Keys(txt='Chrome')
            # 断言
            elif 'assert' in values[1]:
                # 通过反射来执行
                status = getattr(keys, values[1])(**data)
                # 基于断言结果True or False来进行写入的操作
                if status:
                    # 获取assert所在行和列
                    sheet_temp.cell(row=values[0] + 2, column=7).value = 'Pass'
                    # 表格对应的背景颜色
                    sheet_temp.cell(row=values[0] + 2, column=7).fill = PatternFill('solid', fgColor='00b050')
                    # 字体加粗
                    sheet_temp.cell(row=values[0] + 2, column=7).font = Font(bold=True)
                else:
                    sheet_temp.cell(row=values[0] + 2, column=7).value = 'Failed'
                    sheet_temp.cell(row=values[0] + 2, column=7).fill = PatternFill('solid', fgColor='ff0000')
                    sheet_temp.cell(row=values[0] + 2, column=7).font = Font(bold=True)
            # 常规操作
            else:
                getattr(keys, values[1])(**data)
                # keys.input(name=xxx,value=aaaa,txt=dasdsad)
    excel.save(pwd_excel)
