import openpyxl
import os
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment

cur_dir = os.path.dirname(os.path.abspath(__file__))
excel_dir = os.path.join(cur_dir, 'cases.xlsx')
# 加载excel
excel = openpyxl.load_workbook(excel_dir)
# 加载sheet表
sh = excel['Sheet1']
# 解决第三方库没有代码提示
assert isinstance(sh, openpyxl.workbook.workbook.Worksheet)
# 读取指定表格内容
a1 = sh.cell(1, 1).value
print(a1)
# 读取所有内容
value = sh.values
for i in value:
    print(i[0])  # 因为是数组，通过修改角标来读取列
# 写入excel
sh.cell(4, 4).value = '水杯'
# 添加背景颜色
sh.cell(4, 4).fill = PatternFill('solid', fgColor='00b050')
# 字体加粗
sh.cell(4, 4).font = Font(bold=True)

excel.save(excel_dir)
