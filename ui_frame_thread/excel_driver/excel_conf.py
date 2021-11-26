'''
    excel配置
'''
from openpyxl.styles import Font, PatternFill


# Pass配置
def pass_(cell, row, colum):
    # 获取assert所在行和列
    cell(row=row, column=colum).value = 'Pass'
    # 表格对应的背景颜色
    cell(row=row, column=colum).fill = PatternFill('solid', fgColor='00b050')
    # 字体加粗
    cell(row=row, column=colum).font = Font(bold=True)


# Failed配置
def failed_(cell, row, colum):
    # 获取assert所在行和列
    cell(row=row, column=colum).value = 'Failed'
    # 表格对应的背景颜色
    cell(row=row, column=colum).fill = PatternFill('solid', fgColor='ff0000')
    # 字体加粗
    cell(row=row, column=colum).font = Font(bold=True)
