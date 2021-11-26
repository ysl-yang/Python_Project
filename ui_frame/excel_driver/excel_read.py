'''
    excel驱动类
'''
import openpyxl
import logging.config
from ui_frame.excel_driver import excel_conf
from ui_frame.my_conf import log_conf
from ui_frame.ui_keys.Web_Keys import WebKey


# excel驱动类
def excel_runner(path, log):
    # log = logging.config.fileConfig('../my_conf/log.ini')
    # 调用log_conf方法，传入路径
    # log_conf.get_log('../my_conf/log.ini')
    # 创建日志器
    # log = logging.getLogger()
    try:
        # pwd_dir = os.path.dirname(os.path.abspath(__file__))
        # pwd_excel = os.path.join(pwd_dir, 'at_cases.xlsx')
        # 读取excel
        # excel = openpyxl.load_workbook(pwd_excel)
        # excel = openpyxl.load_workbook('./data/cases.xlsx')
        excel = openpyxl.load_workbook(path)
        # 读取sheet表
        # sh = excel['Sheet1']
        # 读取所有sheet
        sheets = excel.sheetnames  # 获取所有的sheet表名
        for sheet in sheets:  # 遍历表名，将遍历的sheet传给excel[sheet]
            sheet_temp = excel[sheet]
            # 解决第三方库没有代码提示
            assert isinstance(sheet_temp, openpyxl.workbook.workbook.Worksheet)
            log.info('----------{}----------'.format(sheet))
            for values in sheet_temp.values:  # 将得到的所有sheet加载
                # 只需要获取用例，不需要标题，通过if判断
                if type(values[0]) is int:  # 从角标为0开始判断是否为数字
                    log.info('正在执行：{}'.format(values[5]))
                    '''
                    用例结构
                        1.编号：不用管
                        2.调用的关键字：结合关键字驱动类应用
                        3.关键字对应的参数：测试数据，结合关键字驱动类应该
                        4.本条用例的行为描述：行为记录，可以添加到日志中
                    '''
                    # 提取数据，name\value\txt是对应的关键字类的输入
                    data = {}
                    data['name'] = values[2]
                    data['value'] = values[3]
                    data['txt'] = values[4]
                    data['expect'] = values[6]
                    # 优化测试数据内容，将所有未None的数据全部从data中清除
                    for key in list(data.keys()):
                        if data[key] is None:
                            del data[key]
                    # print(data)
                    # 基于操作行为和对应参数来执行自动化操作
                    '''
                        用例的操作行为主要分为：
                            1. 实例化，通过一个操作行为实例化关键字驱动类对象
                            2. 常规操作，通过调用已实例化的对象，执行对应的函数。
                            3. 断言操作，判断预期与实际是否符合，将结果填入测试用例中。
                    '''
                    # 实例化关键字驱动
                    if values[1] == 'open_browser':
                        # keys = WebKey(**data)
                        # 获取角标4对应的列，也就是excel中的"输入文本"列
                        keys = WebKey(values[4], log)
                        # Keys(txt='Chrome')
                    # 断言，判断角标1对应的列中是否包含asser的文字，如果有则认为是断言
                    elif 'assert' in values[1]:
                        # 通过反射来执行
                        status = getattr(keys, values[1])(**data)
                        # 基于断言结果True or False来进行写入的操作
                        if status:
                            # 调用excel_conf封装的方法(位置、表格颜色、字体风格)
                            excel_conf.pass_(sheet_temp.cell, values[0] + 2, 8)
                        else:
                            excel_conf.failed_(sheet_temp.cell, values[0] + 2, 8)
                        # excel.save('./data/cases.xlsx')
                        excel.save(path)
                    # 常规操作
                    else:
                        getattr(keys, values[1])(**data)
                        # keys.input(name=xxx,value=aaaa,txt=dasdsad)
    except Exception as e:
        log.exception('运行异常：{}'.format(e))
    finally:
        # 关闭excel
        excel.close()


if __name__ == '__main__':
    excel_runner()
